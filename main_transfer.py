from openpyxl import load_workbook
def main_transfer(kiso_workbook,paste_workbook,btn_value, file_path):
  for y in range(0,11):
    paste_sheet = paste_workbook.worksheets[y]
    if btn_value == "B":
      kiso_sheet = kiso_workbook.worksheets[y]
    
    if btn_value == "A":
      if y <= 9:
        paste_sheet.cell(row=6, column=7).value = None
        paste_sheet.cell(row=6, column=9).value = None
        paste_sheet.cell(row=6, column=12).value = None
        paste_sheet.cell(row=6, column=19).value = paste_sheet.cell(row=6, column=5).value
        paste_sheet.cell(row=6, column=5).value = None
        for row in range(11,19):
          for col in range(6,11):
            paste_sheet.cell(row=row, column=col+13).value = paste_sheet.cell(row=row, column=col).value
            paste_sheet.cell(row=row, column=col).value = None
        for row in range(11,19):
          for col in range(12,14):
            paste_sheet.cell(row=row, column=col+13).value = paste_sheet.cell(row=row, column=col).value
            paste_sheet.cell(row=row, column=col).value = None
      elif y == 10:
        for row_group in [45,72,99,126,153,180,207,234,261,290]:
          paste_sheet.row_dimensions.group(row_group, row_group+18, hidden=True)
    if btn_value == "B":
      if y <= 9:
        paste_sheet.cell(row=6, column=7).value = None
        paste_sheet.cell(row=6, column=9).value = None
        paste_sheet.cell(row=6, column=12).value = None
        paste_sheet.cell(row=6, column=19).value = kiso_sheet.cell(row=6, column=5).value
        paste_sheet.cell(row=6, column=5).value = None
        for row in range(11,19):
          for col in range(6,11):
            paste_sheet.cell(row=row, column=col+13).value = paste_sheet.cell(row=row, column=col).value
            paste_sheet.cell(row=row, column=col).value = None
        for row in range(11,19):
          for col in range(12,14):
            paste_sheet.cell(row=row, column=col+13).value = paste_sheet.cell(row=row, column=col).value
            paste_sheet.cell(row=row, column=col).value = None
        for row in range(19, 38):
          for col in range(6,11):
            paste_sheet.cell(row=row, column=col+13).value = kiso_sheet.cell(row=row, column=col).value
        for row in range(19, 38):
          for col in range(12,14):
            paste_sheet.cell(row=row, column=col+13).value = kiso_sheet.cell(row=row, column=col).value
      elif y == 10:
        for row_group in [45,72,99,126,153,180,207,234,261,290]:
          paste_sheet.row_dimensions.group(row_group, row_group+18, hidden=False)
    if btn_value == "C":
      if y <= 9:
        paste_sheet.cell(row=6, column=7).value = None
        paste_sheet.cell(row=6, column=9).value = None
        paste_sheet.cell(row=6, column=12).value = None
        paste_sheet.cell(row=6, column=19).value = paste_sheet.cell(row=6, column=5).value
        paste_sheet.cell(row=6, column=5).value = None
        for row in range(11,19):
          for col in range(6,11):
            paste_sheet.cell(row=row, column=col+13).value = paste_sheet.cell(row=row, column=col).value
            paste_sheet.cell(row=row, column=col).value = None
        for row in range(11,19):
          for col in range(12,14):
            paste_sheet.cell(row=row, column=col+13).value = paste_sheet.cell(row=row, column=col).value
            paste_sheet.cell(row=row, column=col).value = None
        for row in range(19, 38):
          for col in range(6,11):
            paste_sheet.cell(row=row, column=col+13).value = None
        for row in range(19, 38):
          for col in range(12,14):
            paste_sheet.cell(row=row, column=col+13).value = None
      elif y == 10:
        for row_group in [45,72,99,126,153,180,207,234,261,290]:
          paste_sheet.row_dimensions.group(row_group, row_group+18, hidden=True)
    if btn_value == "D":
      if y <= 9:
        paste_sheet.cell(row=6, column=8).value = None
        paste_sheet.cell(row=6, column=10).value = None
        paste_sheet.cell(row=6, column=14).value = None
        paste_sheet.cell(row=6, column=24).value = paste_sheet.cell(row=6, column=5).value
        paste_sheet.cell(row=6, column=5).value = None
        for row in range(10,29):
          for col in range(8,14):
            paste_sheet.cell(row=row, column=col+16).value = paste_sheet.cell(row=row, column=col).value
            paste_sheet.cell(row=row, column=col).value = None
  if btn_value == "B":
    kiso_workbook.close()
  paste_workbook.save(file_path)
  paste_workbook.close()
