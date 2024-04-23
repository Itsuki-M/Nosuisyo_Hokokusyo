from openpyxl import load_workbook
import os
from tkinter import filedialog, messagebox
from check_survey_day import check_survey_day
from main_transfer import main_transfer

def button2_clicked(btn_value, pre_week, this_week):
  iDir = os.path.abspath(os.path.dirname(__file__))
  folderPath = filedialog.askdirectory(initialdir=iDir)
  folders = ['1.北海道', '2.東北', '3.関東','4.北陸', '5.東海', '6.近畿', '7.中四国', '8.九州', '9.沖縄']
  
  if btn_value == "B":
    nowMonth = this_week
    kiso_Month = int(nowMonth) - 89 if nowMonth[-2:] == "01" else int(nowMonth) - 1
    kiso_day = str(kiso_Month)
  
  for i in range(9):
    if btn_value == "A":
      former_folder = os.path.join(folderPath, '野菜', pre_week, folders[i])
      now_folder = os.path.join(folderPath, '野菜', this_week, folders[i])
    elif btn_value == "B":
      former_folder = os.path.join(folderPath, '野菜', pre_week, folders[i])
      now_folder = os.path.join(folderPath, '基礎', this_week, folders[i])
      kiso_folder = os.path.join(folderPath, '基礎', kiso_day, folders[i])
    elif btn_value == "C":
      former_folder = os.path.join(folderPath, '基礎', pre_week, folders[i])
      now_folder = os.path.join(folderPath, '野菜', this_week, folders[i])
    elif btn_value == "D":
      former_folder = os.path.join(folderPath, '加工食品', pre_week, folders[i])
      now_folder = os.path.join(folderPath, '加工食品', this_week, folders[i])
      
    former_files = [file for file in os.listdir(former_folder) if file.endswith('.xlsx')]
    for former_file in former_files:
      if btn_value == "A" or btn_value == "D":
        now_file = former_file.replace(pre_week, this_week)
      elif btn_value == "B":
        former_file_name = pre_week + 'の週_野菜'
        now_file_name = this_week + '_基礎'
        now_file = former_file.replace(former_file_name, now_file_name)
      elif btn_value == "C":
        former_file_name = pre_week + '_基礎'
        now_file_name = this_week + 'の週_野菜'
        now_file = former_file.replace(former_file_name, now_file_name)
      former_file_path = os.path.join(former_folder, former_file)
      now_file_path = os.path.join(now_folder, now_file)
      
      wb = load_workbook(former_file_path)
      wb.save(now_file_path)
      if btn_value == "B":
        kiso_file = now_file.replace(this_week, kiso_day)
        kiso_file_path = os.path.join(kiso_folder, kiso_file)
        kiso_wb = load_workbook(kiso_file_path)
        main_transfer(kiso_wb, wb, btn_value)
      else:
        main_transfer(None, wb, btn_value, now_file_path)
    
    
  messagebox.showinfo('成功', '報告書を作成しました')